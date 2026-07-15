"""Read/write the uploaded directory sheet (.xlsx or .csv).

The real input is not a flat CSV. Duval_County_Middle_Schools_Directory.xlsx has
its header on row 5 (four preamble rows above it), an Excel Table object, and a
second "Sources & Notes" sheet. So:

  * the header row is DETECTED, never assumed to be row 1;
  * writes go through openpyxl on the ORIGINAL workbook so styling, preamble and
    the second sheet survive, with the Table's ref widened to cover new columns;
  * every write is atomic (tmp + os.replace) and idempotent (a column is added
    only if absent) -- the same discipline the DD-Scrapper enrichers use.
"""

import csv
import os
import shutil

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

NAME_COL = "School Name"
ADDRESS_COL = "Address"
WEBSITE_COL = "Official Website"

GOFAN_LOGO_COL = "gofan logo url"
GOFAN_URL_COL = "gofan url"
WEBSITE_LOGO_COL = "official website logo"

NEW_COLUMNS = (GOFAN_LOGO_COL, GOFAN_URL_COL, WEBSITE_LOGO_COL)

HEADER_SCAN_ROWS = 25


class SheetError(ValueError):
    """Input is not a directory sheet we can process."""


def _norm(v):
    return str(v).strip() if v is not None else ""


def _find_header(matrix):
    """Index of the row containing the School Name header.

    Row 5 in the Duval sheet. Matching is case-insensitive so 'SCHOOL NAME' works.
    """
    for i, row in enumerate(matrix[:HEADER_SCAN_ROWS]):
        lowered = [_norm(c).lower() for c in row]
        if NAME_COL.lower() in lowered:
            return i
    raise SheetError(
        f"Could not find a header row containing {NAME_COL!r} in the first "
        f"{HEADER_SCAN_ROWS} rows. Is this a school directory sheet?"
    )


def read_sheet(path):
    """-> dict(rows, header, header_row_index, sheet_name, ext)

    rows: list of dicts keyed by header name, in file order.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    if ext == ".csv":
        return _read_csv(path)
    raise SheetError(f"Unsupported file type {ext!r}. Upload a .xlsx or .csv file.")


def _read_xlsx(path):
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    matrix = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr_i = _find_header(matrix)
    header = [_norm(c) for c in matrix[hdr_i]]

    rows = []
    for r in matrix[hdr_i + 1 :]:
        rec = {header[j]: _norm(v) for j, v in enumerate(r) if j < len(header) and header[j]}
        if rec.get(NAME_COL):
            rows.append(rec)
    wb.close()
    _require_columns(header)
    return {
        "rows": rows,
        "header": header,
        "header_row_index": hdr_i,
        "sheet_name": ws.title,
        "ext": ".xlsx",
    }


def _read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        matrix = [r for r in csv.reader(fh)]
    if not matrix:
        raise SheetError("The uploaded CSV is empty.")
    hdr_i = _find_header(matrix)
    header = [_norm(c) for c in matrix[hdr_i]]
    rows = []
    for r in matrix[hdr_i + 1 :]:
        rec = {header[j]: _norm(v) for j, v in enumerate(r) if j < len(header) and header[j]}
        if rec.get(NAME_COL):
            rows.append(rec)
    _require_columns(header)
    return {
        "rows": rows,
        "header": header,
        "header_row_index": hdr_i,
        "sheet_name": None,
        "ext": ".csv",
    }


def _require_columns(header):
    lowered = {h.lower() for h in header}
    missing = [c for c in (NAME_COL, ADDRESS_COL) if c.lower() not in lowered]
    if missing:
        raise SheetError(f"Sheet is missing required column(s): {', '.join(missing)}")


def write_csv(out_path, header, rows):
    """Flat CSV of header + data rows. Atomic."""
    fields = list(header) + [c for c in NEW_COLUMNS if c not in header]
    tmp = out_path + ".tmp"
    with open(tmp, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})
    os.replace(tmp, out_path)
    return out_path


def write_xlsx(src_path, out_path, meta, rows):
    """Copy the original workbook and append the new columns in place.

    Editing a copy of the original (rather than building a fresh workbook) is what
    keeps the preamble rows, cell styling, and the "Sources & Notes" sheet intact.
    """
    # Keep the .xlsx suffix on the temp file: openpyxl refuses to open a path
    # whose extension it doesn't recognise, so out_path + ".tmp" would fail.
    tmp = out_path + ".tmp.xlsx"
    shutil.copyfile(src_path, tmp)

    wb = load_workbook(tmp)
    ws = wb[meta["sheet_name"]] if meta["sheet_name"] in wb.sheetnames else wb.worksheets[0]

    hdr_row = meta["header_row_index"] + 1  # openpyxl is 1-indexed
    header = list(meta["header"])

    # Idempotent: reuse a column if a previous run already added it.
    col_of = {}
    for name in NEW_COLUMNS:
        found = None
        for j, h in enumerate(header):
            if h.strip().lower() == name.lower():
                found = j + 1
                break
        if found is None:
            found = len(header) + 1
            header.append(name)
            ws.cell(row=hdr_row, column=found, value=name)
        col_of[name] = found

    # Mirror the header row's style onto any newly created header cells so the
    # appended columns don't look bolted on.
    template = ws.cell(row=hdr_row, column=1)
    for name in NEW_COLUMNS:
        c = ws.cell(row=hdr_row, column=col_of[name])
        if c.value == name and template.has_style:
            c._style = template._style

    for i, rec in enumerate(rows):
        excel_row = hdr_row + 1 + i
        for name in NEW_COLUMNS:
            ws.cell(row=excel_row, column=col_of[name], value=rec.get(name, "") or "")

    _widen_tables(ws, len(header))

    for name in NEW_COLUMNS:
        ws.column_dimensions[get_column_letter(col_of[name])].width = 52

    wb.save(tmp)
    wb.close()
    os.replace(tmp, out_path)
    return out_path


def _widen_tables(ws, ncols):
    """Extend any Excel Table ref to cover the appended columns.

    Without this the styled table stops short and the new columns render outside
    it. openpyxl exposes tables as a name -> ref mapping.
    """
    tables = getattr(ws, "tables", None)
    if not tables:
        return
    for name in list(tables.keys()):
        tbl = tables[name]
        try:
            start, end = tbl.ref.split(":")
            start_row = int("".join(ch for ch in start if ch.isdigit()))
            end_row = int("".join(ch for ch in end if ch.isdigit()))
            start_col = "".join(ch for ch in start if ch.isalpha())
            tbl.ref = f"{start_col}{start_row}:{get_column_letter(ncols)}{end_row}"
        except Exception:
            # A malformed ref must not cost the user their results.
            continue
